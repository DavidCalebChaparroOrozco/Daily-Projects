# Importing necessary libraries
import pandas as pd
import random
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Generate dummy data
names = ["David Caleb", "Andrea López", "Carlos Mendoza", "Sofía Ramírez", "Jorge Castillo"]
clients = ["Tech Solutions", "Global Market", "Innova Corp", "Future Vision", "SecureData Ltd"]
dates = pd.date_range(start="2024-01-01", periods=30, freq="D")
hours_worked = [round(random.uniform(3, 10), 2) for _ in range(30)]

data = {
    "Date": [date.strftime("%Y-%m-%d") for date in dates],
    "Employee": [random.choice(names) for _ in range(30)],
    "Client": [random.choice(clients) for _ in range(30)],
    "Hours Worked": hours_worked
}

# Create DataFrame
df = pd.DataFrame(data)

# Create an Excel file with professional formatting
wb = Workbook()
ws = wb.active
ws.title = "Work Hours Report"

# Add headers
headers = ["Date", "Employee", "Client", "Hours Worked"]
ws.append(headers)

# Header styles
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
header_alignment = Alignment(horizontal="center", vertical="center")

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# Add data with formatting
border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

for row in df.itertuples(index=False):
    ws.append(row)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=ws.max_row, column=col_num)
        cell.border = border
        if col_num == 4:  # Center the hours column
            cell.alignment = Alignment(horizontal="center")

# Adjust column widths
for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = max_length + 2

# Save file
file_path = "data/Work_Hours_Report.xlsx"
wb.save(file_path)

file_path